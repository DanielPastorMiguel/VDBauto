import configuracion
import utiles
from openpyxl import Workbook
from openpyxl.styles  import PatternFill, Font, Border, Side, Alignment

class Vdb_generado(object):
    """Clase que tiene como atributo el fichero VDB que genera la aplicacion.
    Esta implementada siguiendo el patron Singleton para evitar multiples instancias de esta clase"""

    workbook = None
    hoja = None
    hoja2 = None

    ruta = None
    conf = None

    borde_cabecera = None
    alineamiento_cabecera = None
    alineamiento_celdas_insertar = None
    patternFill_cabecera = None
    fuente_cabecera = None
    nombre_hoja_info_extra = "Info Extra"

    def __new__(cls):
        if not hasattr(cls, 'instance'):
          cls.instance = super(Vdb_generado, cls).__new__(cls)
        return cls.instance

    def __init__(self):
        self.conf = configuracion.Configuracion()
        self.ruta = self.conf.ruta_vdb_generado+"_"+utiles.get_fecha_hora_actual()+".xlsx"
        self.inicializar_valores_cabecera()
        self.crear_vdb()
        self.conf_vdb()
        self.conf_hoja2()
    
    def inicializar_valores_cabecera(self):
        self.patternFill_cabecera = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.fuente_cabecera = Font(color="FFFFFF")
        self.alineamiento_cabecera = Alignment(
                horizontal='center', 
                vertical='center',
                text_rotation=0,
                wrap_text=False,
                shrink_to_fit=False,
                indent=0)
        self.borde_cabecera = Border(
                left=Side(border_style="thin", color='000000'),
                right=Side(border_style="thin", color='000000'),
                top=Side(border_style="thin", color='000000'),
                bottom=Side(border_style="thick", color='000000')
            )
    
    def crear_vdb(self):
        self.workbook = Workbook()
        self.hoja = self.workbook.active
        self.workbook.create_sheet(self.nombre_hoja_info_extra)
        self.hoja2 = self.workbook[self.nombre_hoja_info_extra]
    
    def conf_vdb(self):
        cabecera_vdb = utiles.get_cabecera_xlsx(self.conf.vdb_hoja)
        self.hoja.append(cabecera_vdb[0]) #insertamos la cabecera del vbd introducido al vbd generado
        self.hoja.append(cabecera_vdb[1]) #la cabecera consta en 2 filas
        self.formato_cabecera()
        self.combinar_celdas()
        self.ancho_columnas_criticidades()
        self.workbook.save(self.ruta)
    
    def conf_hoja2(self):
        cabecera = ["VULNERABILIDADES ERROR", "", "POSIBLES CVE REJECTED", "", "VULNERABILIDADES PUEDEN CERRAR", "", "VULNERABILIDADES CONTINUAN"]
        self.hoja2.append(cabecera)
        celda_vul_error = self.hoja2.cell(row=1, column=1)
        celda_vul_rejected = self.hoja2.cell(row=1, column=3)
        celda_vul_cerrar = self.hoja2.cell(row=1, column=5)
        celda_vul_continuan = self.hoja2.cell(row=1, column=7)

        celda_vul_error.fill = self.patternFill_cabecera
        celda_vul_error.font = self.fuente_cabecera
        celda_vul_error.alignment = self.alineamiento_cabecera
        celda_vul_error.border = self.borde_cabecera

        celda_vul_rejected.fill = self.patternFill_cabecera
        celda_vul_rejected.font = self.fuente_cabecera
        celda_vul_rejected.alignment = self.alineamiento_cabecera
        celda_vul_rejected.border = self.borde_cabecera

        celda_vul_cerrar.fill = self.patternFill_cabecera
        celda_vul_cerrar.font = self.fuente_cabecera
        celda_vul_cerrar.alignment = self.alineamiento_cabecera
        celda_vul_cerrar.border = self.borde_cabecera

        celda_vul_continuan.fill = self.patternFill_cabecera
        celda_vul_continuan.font = self.fuente_cabecera
        celda_vul_continuan.alignment = self.alineamiento_cabecera
        celda_vul_continuan.border = self.borde_cabecera

        self.hoja2.column_dimensions[celda_vul_error.column_letter].width = len(celda_vul_error.value)+5
        self.hoja2.column_dimensions[celda_vul_rejected.column_letter].width = len(celda_vul_rejected.value)+5
        self.hoja2.column_dimensions[celda_vul_cerrar.column_letter].width = len(celda_vul_cerrar.value)+5
        self.hoja2.column_dimensions[celda_vul_continuan.column_letter].width = len(celda_vul_continuan.value)+5

    def formato_cabecera(self):
        for indice_columna in range(1, self.hoja.max_column+1):
            celda_superior= self.hoja.cell(row=1, column=indice_columna) #celda de la fila1
            celda_inferior = self.hoja.cell(row=2, column=indice_columna) #celda de la fila2

            patron = self.patternFill_cabecera
            fuente = self.fuente_cabecera
            alineado = self.alineamiento_cabecera

            celda_superior.fill = patron
            celda_inferior.fill = patron
            celda_superior.font = fuente
            celda_inferior.font = fuente
            celda_superior.border = self.borde_cabecera
            celda_inferior.border = self.borde_cabecera
            celda_superior.alignment = alineado
            celda_inferior.alignment = alineado

            valor = celda_superior.value
            if valor == None: valor = "example" #para dar una longitud predefinida a las columnas que tengan vacia la celda de la cabecera
            
            self.hoja.column_dimensions[celda_superior.column_letter].width = len(valor)+2

    def combinar_celdas(self):
        for indice_columna in range(1, self.hoja.max_column+1):
            celda_inferior = self.hoja.cell(row=2, column=indice_columna) #celda de la fila2
            if celda_inferior.value == None:
                self.hoja.merge_cells(start_row=1, start_column=indice_columna, end_row=2, end_column=indice_columna) #combinamos las celdas de las dos primeras filas para esta columna

        indices_combinar_horizontal = {} #almacena ej: 8:[9,10], 11:[12] Donde 8, 9 y 10, y 11 y 12 son filas que tienen que combinarse
        indice_max = self.hoja.max_column+1
        indice_columna = 1
        indice_actual = indice_columna
        while indice_columna < indice_max-1:
            celda_superior = self.hoja.cell(row=1, column=indice_columna+1) #celda de la fila1
            if celda_superior.value == None:
                try:
                    indices_combinar_horizontal[indice_actual]
                    indices_combinar_horizontal[indice_actual].append(indice_columna+1)
                except KeyError:
                    indices_combinar_horizontal[indice_actual] = [indice_columna+1]
                indice_columna+=1
            else:
                indice_columna+=1
                indice_actual = indice_columna
            
        for clave, valor in indices_combinar_horizontal.items():
            self.hoja.merge_cells(start_row=1, start_column=clave, end_row=1, end_column=valor[len(valor)-1]) #combina la primera con la ultima

    def ancho_columnas_criticidades(self):
        for clave, valor in self.conf.configuracion["vdb_generado"]["columnas_criticidades"].items():
            if clave != "fecha_publicacion": self.hoja.column_dimensions[valor].width = 10
        
    def insertar_celda(self, hoja, fila, columna, valor):
        celda = hoja.cell(row=fila, column=columna)
        celda.value = valor
        celda.alignment = self.alineamiento_cabecera
        self.workbook.save(self.ruta)

    def insertar_fila_vdb_generado(self, fila):
        self.hoja.append(fila)
        for celda in self.hoja[self.hoja.max_row]:
            celda.alignment = self.alineamiento_cabecera
        self.workbook.save(self.ruta)